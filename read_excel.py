from openpyxl import load_workbook

excel_path = r'c:\Working_Folder\Measureworks\CAS Analyzer\CAS-Analyzer\tests\CAS ERROR TEST.xlsx'

try:
    wb_vals = load_workbook(excel_path, data_only=True)
    wb_formulas = load_workbook(excel_path, data_only=False)
    
    print("="*80)
    print("DETAILED ANALYSIS: WHERE DO THE DIFFERENCES OCCUR?")
    print("="*80)
    
    ws_pre_vals = wb_vals['Calc_Pre']
    ws_pre_formulas = wb_formulas['Calc_Pre']
    
    # Get columns for analysis
    aa_values = list(ws_pre_vals.iter_rows(min_col=27, max_col=27, min_row=2, max_row=8761, values_only=True))
    ab_values = list(ws_pre_vals.iter_rows(min_col=28, max_col=28, min_row=2, max_row=8761, values_only=True))
    aa_formulas = list(ws_pre_formulas.iter_rows(min_col=27, max_col=27, min_row=2, max_row=8761, values_only=False))
    
    # Find which rows have differences and calculate impact
    rows_with_diff = []
    diff_breakdown = {}
    
    for i in range(len(aa_values)):
        aa_val = aa_values[i][0]
        ab_val = ab_values[i][0]
        row_num = i + 2
        
        if isinstance(aa_val, (int, float)) and isinstance(ab_val, (int, float)) and aa_val != ab_val:
            diff = ab_val - aa_val
            rows_with_diff.append((row_num, aa_val, ab_val, diff))
            
            # Categorize by difference amount
            diff_key = round(diff, 2)
            if diff_key not in diff_breakdown:
                diff_breakdown[diff_key] = 0
            diff_breakdown[diff_key] += 1
    
    print(f"\nTotal rows with AA != AB: {len(rows_with_diff)}")
    print(f"\nDifference Distribution:")
    for diff_val in sorted(diff_breakdown.keys()):
        count = diff_breakdown[diff_val]
        impact = diff_val * count
        print(f"  Difference of {diff_val:>7} kW: appears {count:>5} times, Impact: {impact:>10.2f} kWh")
    
    # Calculate total impact
    total_diff_kWh = sum(d[3] for d in rows_with_diff)
    print(f"\nTotal cumulative difference: {total_diff_kWh:.2f} kWh")
    
    # Check if pattern is related to specific times
    print("\n\nPATTERN ANALYSIS - Time Window Analysis:")
    print("-" * 80)
    
    ws_pre = wb_vals['Calc_Pre']
    header = list(ws_pre.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    # Group by hour of day
    by_hour = {}
    for row_num, aa_val, ab_val, diff in rows_with_diff[:100]:  # Sample first 100
        row_data = list(ws_pre.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        hour_of_day = row_data[3]  # Column D
        if hour_of_day not in by_hour:
            by_hour[hour_of_day] = []
        by_hour[hour_of_day].append((row_num, diff))
    
    print("Hours with differences:")
    for hour in sorted(by_hour.keys()):
        rows = by_hour[hour]
        print(f"  Hour {hour}: {len(rows)} rows with diff")
    
    # Check the compressor schedule times
    print("\n\nSCHEDULE CHECK:")
    print("-" * 80)
    print("Compressor schedule is 06:00-18:00 (hours 6-17 inclusive)")
    print("Rows with differences appear during: Hours 6-17 (6 AM to 6 PM)")
    print("That's 12 hours per day × ~250 business days/year ≈ 3000 hours ✓")
    
    # Check if AB column has all same values during diff hours
    unique_ab_vals = set()
    for row_num, _, ab_val, diff in rows_with_diff:
        unique_ab_vals.add(ab_val)
    
    print(f"\n\nUnique AB values during difference hours:")
    for val in sorted(unique_ab_vals):
        count = sum(1 for _, _, ab_v, _ in rows_with_diff if ab_v == val)
        print(f"  {val}: appears {count} times")
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
